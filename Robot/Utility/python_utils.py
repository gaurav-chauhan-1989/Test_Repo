import openpyxl
import yaml

class Utils:

    def read_excel_data(self, filename, sheet):
        wb = openpyxl.load_workbook(filename)
        sh = wb[sheet]
        data = []
        for i in range(2, sh.max_row+1):
            row_data = []
            for j in range(1, sh.max_column+1):
                row_data.append(sh.cell(row=i, column=j).value)
            data.append(row_data)
        return data

    @staticmethod
    def yaml_parser(filename):
        with open(filename, 'r') as file:
            return yaml.safe_load(file)
