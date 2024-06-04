import openpyxl

'''wb = openpyxl.load_workbook("E:\Sheet\wrt.xlsx")
sh = wb.active
a=[]
for i in range(1, sh.max_row+1):
    for j in range(1, sh.max_column+1):
        a.append(sh.cell(row=i, column=j).value)
print(a)'''


def read_excel_data(filename, sheet):
    wb = openpyxl.load_workbook(filename)
    sh = wb[sheet]
    data = []
    for i in range(2, sh.max_row + 1):
        row_data = []
        for j in range(1, sh.max_column + 1):
            row_data.append(sh.cell(row=i, column=j).value)
        data.append(row_data)
    return data


d = read_excel_data('E:\\Sheet\\sht.xlsx', 'Sheet')
print(d)