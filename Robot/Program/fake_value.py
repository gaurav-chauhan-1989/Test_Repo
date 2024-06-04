from faker import Faker
from openpyxl import Workbook

wb = Workbook()
sh = wb.active
fake_data = Faker()
for i in range(1,11):
    sh.cell(row=i, column=1).value = fake_data.phone_number()
    sh.cell(row=i, column=2).value = fake_data.name()
    sh.cell(row=i, column=3).value = fake_data.email()
    sh.cell(row=i, column=4).value = fake_data.address()
wb.save("E:\Sheet\data.xlsx")


